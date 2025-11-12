"""
Adapter FastAPI para orquestração do robô de comissões.

Este adapter apenas orquestra chamadas ao robô Python existente,
sem alterar a lógica de negócio.
"""

import os
import json
import subprocess
import uuid
import asyncio
import sys
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, PlainTextResponse
from pydantic import BaseModel
import aiofiles
from dotenv import load_dotenv

# Carregar variáveis de ambiente
# Carregar .env do diretório do adapter
adapter_dir = Path(__file__).parent
env_path = adapter_dir / ".env"
load_dotenv(dotenv_path=env_path, override=True)

# Configuração
# Tentar obter ROBO_ROOT_PATH do .env, se não existir usar diretório pai do adapter (raiz do projeto)
default_path = (
    adapter_dir.parent.parent.resolve()
    if adapter_dir.name == "adapter"
    else Path(os.getcwd())
)
ROBO_ROOT_PATH = os.getenv("ROBO_ROOT_PATH")
if not ROBO_ROOT_PATH:
    # Fallback: usar caminho padrão baseado na estrutura do projeto
    ROBO_ROOT_PATH = str(default_path)
else:
    # Garantir caminho absoluto
    ROBO_ROOT_PATH = str(Path(ROBO_ROOT_PATH).resolve())

# Garantir que o caminho raiz do robô está no sys.path para imports diretos
try:
    if ROBO_ROOT_PATH not in sys.path:
        sys.path.insert(0, ROBO_ROOT_PATH)
except Exception:
    pass

PROGRESS_FILE = os.path.join(ROBO_ROOT_PATH, "progress.json")

# ==================== LOGGING (Arquivo) ====================
import logging
from logging.handlers import RotatingFileHandler

LOG_FILE = os.path.join(ROBO_ROOT_PATH, "adapter.log")

try:
    _root_logger = logging.getLogger()
    _root_logger.setLevel(logging.INFO)
    has_file_handler = any(
        isinstance(h, RotatingFileHandler)
        and getattr(h, "baseFilename", "").endswith("adapter.log")
        for h in _root_logger.handlers
    )
    if not has_file_handler:
        _fh = RotatingFileHandler(
            LOG_FILE, maxBytes=1_000_000, backupCount=2, encoding="utf-8"
        )
        _fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        _root_logger.addHandler(_fh)
except Exception:
    pass

app = FastAPI(
    title="Adapter Robô de Comissões",
    description="Backend adapter para orquestração do robô de comissões",
    version="1.0.0",
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== MODELS ====================


class BulkApplyRequest(BaseModel):
    escopo: Dict[str, List[str]]  # ex: {"linha": ["A", "B"], "grupo": ["G1"]}
    campos: Dict[str, Any]  # campos a definir
    modo: str  # "criar" ou "atualizar"
    previewOnly: bool = True


class SaveRequest(BaseModel):
    data: List[Dict[str, Any]]
    preserve_columns: bool = True


class ProgressResponse(BaseModel):
    job_id: str
    percent: float
    etapa: str
    mensagens: List[str]
    status: str  # "em_andamento", "concluido", "erro"


# ==================== HELPER FUNCTIONS ====================


def get_regras_path() -> Path:
    """Retorna caminho do arquivo Regras_Comissoes.xlsx"""
    path = Path(ROBO_ROOT_PATH) / "Regras_Comissoes.xlsx"
    # Garantir caminho absoluto
    return path.resolve()


def get_resultado_path() -> Optional[Path]:
    """Retorna caminho do arquivo de resultado mais recente"""
    pattern = "Comissoes_Calculadas_*.xlsx"
    files = list(Path(ROBO_ROOT_PATH).glob(pattern))
    if not files:
        return None
    # Ordenar por data de modificação (mais recente primeiro)
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0]


def read_excel_sheet(filepath: Path, sheet_name: str) -> pd.DataFrame:
    """Lê uma aba do Excel preservando ordem de colunas"""
    try:
        df = pd.read_excel(
            filepath, sheet_name=sheet_name, dtype=str, keep_default_na=False
        )
        # Preservar ordem original das colunas
        return df
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao ler aba {sheet_name}: {str(e)}"
        )


def write_excel_sheet(
    filepath: Path, sheet_name: str, df: pd.DataFrame, preserve_order: bool = True
):
    """Escreve uma aba no Excel preservando ordem de colunas"""
    try:
        # Se arquivo existe, carregar para preservar outras abas
        if filepath.exists():
            with pd.ExcelWriter(
                filepath, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Criar novo arquivo
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao salvar aba {sheet_name}: {str(e)}"
        )


# ==================== ENDPOINTS - REGRAS ====================


@app.get("/regras/abas")
async def listar_abas_regras():
    """Lista todas as abas do arquivo Regras_Comissoes.xlsx"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        # Retornar erro em vez de lista vazia para facilitar debug
        raise HTTPException(
            status_code=404,
            detail=f"Arquivo Regras_Comissoes.xlsx não encontrado em: {regras_path}",
        )

    try:
        wb = load_workbook(regras_path, read_only=True)
        return {"abas": wb.sheetnames}
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao ler arquivo de regras: {str(e)}"
        )


@app.get("/regras/aba/{nome_aba}")
async def ler_aba_regras(
    nome_aba: str,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=1000),
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = Query("asc", regex="^(asc|desc)$"),
    filters: Optional[str] = None,  # JSON string com filtros
    all_pages: bool = Query(False),  # Buscar todas as páginas sem paginação
):
    """Lê uma aba do Regras_Comissoes.xlsx com paginação e filtros"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        raise HTTPException(
            status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado"
        )

    df = read_excel_sheet(regras_path, nome_aba)

    # Aplicar filtros
    if filters:
        try:
            filter_dict = json.loads(filters)
            for col, value in filter_dict.items():
                if col in df.columns and value:
                    df = df[
                        df[col]
                        .astype(str)
                        .str.contains(str(value), case=False, na=False)
                    ]
        except Exception:
            pass

    # Ordenação
    if sort_by and sort_by in df.columns:
        ascending = sort_order == "asc"
        df = df.sort_values(by=sort_by, ascending=ascending)

    # Paginação
    total = len(df)
    if all_pages:
        # Retornar todos os dados sem paginação
        df_page = df
    else:
        start = (page - 1) * size
        end = start + size
        df_page = df.iloc[start:end]

    return {
        "data": df_page.to_dict(orient="records"),
        "total": total,
        "page": page if not all_pages else 1,
        "size": len(df_page) if all_pages else size,
        "columns": list(df.columns),
    }


@app.get("/regras/aba/{nome_aba}/valores-unicos/{coluna}")
async def obter_valores_unicos(nome_aba: str, coluna: str):
    """Retorna valores únicos de uma coluna específica da aba"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        raise HTTPException(
            status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado"
        )

    df = read_excel_sheet(regras_path, nome_aba)

    if coluna not in df.columns:
        raise HTTPException(
            status_code=404,
            detail=f"Coluna '{coluna}' não encontrada na aba '{nome_aba}'",
        )

    # Extrair valores únicos, remover NaN e valores vazios, converter para string e ordenar
    valores_unicos = (
        df[coluna]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", None)
        .dropna()
        .unique()
        .tolist()
    )
    valores_unicos.sort()

    return {"coluna": coluna, "valores": valores_unicos}


@app.post("/regras/aba/{nome_aba}/save")
async def salvar_aba_regras(nome_aba: str, request: SaveRequest):
    """Salva alterações em uma aba do Regras_Comissoes.xlsx"""
    regras_path = get_regras_path()

    # Ler aba atual para preservar colunas
    if regras_path.exists():
        df_existing = read_excel_sheet(regras_path, nome_aba)
        columns_order = list(df_existing.columns)
    else:
        columns_order = list(request.data[0].keys()) if request.data else []

    # Criar DataFrame
    df_new = pd.DataFrame(request.data)

    # Preservar ordem de colunas
    if request.preserve_columns and columns_order:
        # Adicionar colunas faltantes
        for col in columns_order:
            if col not in df_new.columns:
                df_new[col] = ""
        # Reordenar
        df_new = df_new[columns_order]

    # Salvar
    write_excel_sheet(regras_path, nome_aba, df_new, preserve_order=True)

    return {"success": True, "message": f"Aba {nome_aba} salva com sucesso"}


@app.post("/regras/aba/{nome_aba}/apply-bulk")
async def aplicar_massa_regras(nome_aba: str, request: BulkApplyRequest):
    """Aplica alterações em massa na aba"""
    regras_path = get_regras_path()
    if not regras_path.exists():
        raise HTTPException(
            status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado"
        )

    df = read_excel_sheet(regras_path, nome_aba)

    # Aplicar filtros de escopo
    mask = pd.Series([True] * len(df))
    for col, values in request.escopo.items():
        if col in df.columns:
            mask = mask & df[col].isin(values)

    df_filtered = df[mask].copy()

    if request.previewOnly:
        # Pré-visualização
        if request.modo == "criar":
            # Gerar combinações
            # Implementação simplificada - criar novas linhas
            preview_data = []
            for idx, row in df_filtered.iterrows():
                new_row = row.to_dict()
                for key, value in request.campos.items():
                    new_row[key] = value
                preview_data.append(new_row)
        else:
            # Atualizar existentes
            preview_data = df_filtered.copy()
            for key, value in request.campos.items():
                preview_data[key] = value
            preview_data = preview_data.to_dict(orient="records")

        return {
            "preview": preview_data[:100],  # Limitar preview
            "total_afetadas": len(df_filtered),
            "previewOnly": True,
        }
    else:
        # Aplicar alterações
        if request.modo == "criar":
            # Criar novas linhas
            new_rows = []
            for idx, row in df_filtered.iterrows():
                new_row = row.to_dict()
                for key, value in request.campos.items():
                    new_row[key] = value
                new_rows.append(new_row)
            df_new = pd.DataFrame(new_rows)
            df = pd.concat([df, df_new], ignore_index=True)
        else:
            # Atualizar existentes
            for key, value in request.campos.items():
                if key in df.columns:
                    df.loc[mask, key] = value

        write_excel_sheet(regras_path, nome_aba, df, preserve_order=True)
        return {"success": True, "total_afetadas": len(df_filtered)}


# ==================== ENDPOINTS - GERENCIAMENTO DE REGRAS (PESOS_METAS / CONFIG_COMISSAO) ====================


@app.get("/api/regras/pesos-metas")
async def api_get_pesos_metas():
    """Retorna dados da planilha PESOS_METAS como JSON."""
    regras_path = get_regras_path()
    if not Path(regras_path).exists():
        raise HTTPException(
            status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado"
        )

    try:
        df = read_excel_sheet(regras_path, "PESOS_METAS")
        # Normalizar NaN -> ""
        df = df.fillna("")
        return df.to_dict(orient="records")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao ler PESOS_METAS: {str(e)}"
        )


"""
Nota: Em Pydantic v2, modelos root devem usar RootModel.
Para simplificar e evitar dependência de RootModel, aceitamos diretamente
List[Dict[str, Any]] no endpoint abaixo.
"""


@app.post("/api/regras/pesos-metas")
async def api_update_pesos_metas(payload: List[Dict[str, Any]]):
    """Atualiza a planilha PESOS_METAS.
    Validação: soma horizontal dos componentes deve ser ~100.
    """
    regras_path = get_regras_path()
    registros = payload

    if not isinstance(registros, list) or not registros:
        raise HTTPException(status_code=400, detail="Payload inválido")

    df = pd.DataFrame(registros)

    # Componentes padrão conhecidos (presentes ou não)
    componentes = [
        "faturamento_linha",
        "conversao_linha",
        "rentabilidade",
        "faturamento_individual",
        "conversao_individual",
        "retencao_clientes",
        "meta_fornecedor_1",
        "meta_fornecedor_2",
    ]

    # Converter colunas de componentes para numérico quando existirem
    for col in componentes:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Validar soma horizontal ~ 100 (tolerância 0.1)
    if any(col in df.columns for col in componentes):
        presentes = [c for c in componentes if c in df.columns]
        somas = df[presentes].sum(axis=1)
        tolerancia = 0.1
        invalidas = (~((somas >= 100 - tolerancia) & (somas <= 100 + tolerancia))).any()
        if invalidas:
            raise HTTPException(
                status_code=400, detail="Soma de pesos por cargo deve ser 100% (±0.1)"
            )

    try:
        write_excel_sheet(Path(regras_path), "PESOS_METAS", df, preserve_order=True)
        return {"success": True, "message": "PESOS_METAS atualizado."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao salvar PESOS_METAS: {str(e)}"
        )


def _read_config_comissao_df() -> pd.DataFrame:
    regras_path = get_regras_path()
    if not Path(regras_path).exists():
        raise HTTPException(
            status_code=404, detail="Arquivo Regras_Comissoes.xlsx não encontrado"
        )
    return read_excel_sheet(regras_path, "CONFIG_COMISSAO").fillna("")


@app.get("/api/regras/config-comissao/context-options")
async def api_get_config_context_options():
    df = _read_config_comissao_df()

    def uniq(col: str) -> List[str]:
        return (
            sorted(
                pd.Series(df[col])
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .unique()
                .tolist()
            )
            if col in df.columns
            else []
        )

    return {
        "linha": uniq("linha"),
        "grupo": uniq("grupo"),
        "subgrupo": uniq("subgrupo"),
        "tipo_mercadoria": uniq("tipo_mercadoria"),
        "cargo": uniq("cargo"),
    }


@app.post("/api/regras/config-comissao/query")
async def api_query_config_comissao(filters: Dict[str, Any]):
    df = _read_config_comissao_df()
    query = df.copy()
    for key, value in (filters or {}).items():
        if value and value != "Todos" and key in query.columns:
            query = query[query[key].astype(str).str.strip() == str(value).strip()]
    return query.to_dict(orient="records")


@app.put("/api/regras/config-comissao/update-line")
async def api_update_config_comissao_line(rowData: Dict[str, Any]):
    """Atualiza uma linha com base no contexto (chaves) e campos editáveis.
    Campos alvo: taxa_rateio_maximo_pct, fatia_cargo_pct
    """
    df = _read_config_comissao_df()
    regras_path = get_regras_path()

    # Determinar chaves de contexto para identificação
    keys_ctx = ["linha", "grupo", "subgrupo", "tipo_mercadoria", "cargo"]
    mask = pd.Series([True] * len(df))
    for k in keys_ctx:
        if k in rowData and rowData[k] not in (None, "") and k in df.columns:
            mask = mask & (df[k].astype(str).str.strip() == str(rowData[k]).strip())

    if mask.sum() == 0:
        raise HTTPException(
            status_code=404, detail="Nenhuma linha encontrada para atualização"
        )

    # Aplicar atualizações se presentes
    if "taxa_rateio_maximo_pct" in rowData and "taxa_rateio_maximo_pct" in df.columns:
        df.loc[mask, "taxa_rateio_maximo_pct"] = pd.to_numeric(
            rowData["taxa_rateio_maximo_pct"], errors="coerce"
        ).fillna(0)
    if "fatia_cargo_pct" in rowData and "fatia_cargo_pct" in df.columns:
        df.loc[mask, "fatia_cargo_pct"] = pd.to_numeric(
            rowData["fatia_cargo_pct"], errors="coerce"
        ).fillna(0)

    try:
        write_excel_sheet(Path(regras_path), "CONFIG_COMISSAO", df, preserve_order=True)
        return {"success": True, "linhas_atualizadas": int(mask.sum())}
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao salvar CONFIG_COMISSAO: {str(e)}"
        )


class BatchActionItem(BaseModel):
    valor: float


class BatchAction(BaseModel):
    taxa_rateio_maximo_pct: Optional[BatchActionItem] = None
    fatia_cargo_pct: Optional[BatchActionItem] = None


class BatchRequest(BaseModel):
    escopo: Dict[str, Any] = {}
    acao: BatchAction


def _apply_batch_logic(
    df: pd.DataFrame, batch_data: Dict[str, Any]
) -> (pd.DataFrame, int):
    escopo = (batch_data or {}).get("escopo", {})
    acao = (batch_data or {}).get("acao", {})

    query = df.copy()
    for key, value in (escopo or {}).items():
        if value and value != "Todos" and key in query.columns:
            query = query[query[key].astype(str).str.strip() == str(value).strip()]

    idx = query.index
    df_mod = df.copy()

    if isinstance(acao, dict):
        if (
            acao.get("taxa_rateio_maximo_pct") is not None
            and "taxa_rateio_maximo_pct" in df_mod.columns
        ):
            val = float(acao["taxa_rateio_maximo_pct"].get("valor"))
            df_mod.loc[idx, "taxa_rateio_maximo_pct"] = val
        if (
            acao.get("fatia_cargo_pct") is not None
            and "fatia_cargo_pct" in df_mod.columns
        ):
            val = float(acao["fatia_cargo_pct"].get("valor"))
            df_mod.loc[idx, "fatia_cargo_pct"] = val

    return df_mod, len(idx)


@app.post("/api/regras/config-comissao/dry-run")
async def api_dry_run_config_comissao(batch: BatchRequest):
    df = _read_config_comissao_df()
    df_mod, afetadas = _apply_batch_logic(df, batch.dict())
    return {"linhas_afetadas": int(afetadas)}


@app.post("/api/regras/config-comissao/apply-batch")
async def api_apply_batch_config_comissao(batch: BatchRequest):
    df = _read_config_comissao_df()
    regras_path = get_regras_path()
    df_mod, afetadas = _apply_batch_logic(df, batch.dict())
    try:
        write_excel_sheet(
            Path(regras_path), "CONFIG_COMISSAO", df_mod, preserve_order=True
        )
        return {"success": True, "message": f"{afetadas} regras atualizadas."}
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erro ao salvar CONFIG_COMISSAO: {str(e)}"
        )


@app.post("/api/regras/config-comissao/validate-pe")
async def api_validate_config_comissao_pe(contexto: Dict[str, Any]):
    df = _read_config_comissao_df()
    query = df.copy()
    for key, value in (contexto or {}).items():
        if value is None:
            continue
        if key in query.columns:
            if value == "":
                query = query[query[key].astype(str).str.strip() == ""]
            else:
                query = query[query[key].astype(str).str.strip() == str(value).strip()]

    if query.empty:
        return {
            "soma_pe": 0,
            "status": "vazio",
            "message": "Nenhuma regra encontrada para este contexto exato.",
        }

    if "fatia_cargo_pct" not in query.columns:
        raise HTTPException(
            status_code=400, detail="Coluna 'fatia_cargo_pct' não encontrada"
        )

    soma = pd.to_numeric(query["fatia_cargo_pct"], errors="coerce").fillna(0).sum()
    soma = round(float(soma), 2)
    if abs(soma - 100.0) < 0.01:
        return {"soma_pe": soma, "status": "ok", "message": f"Soma correta: {soma}%"}
    else:
        return {
            "soma_pe": soma,
            "status": "erro",
            "message": f"Erro de soma: {soma}%. (Esperado: 100%)",
        }


# ==================== ENDPOINTS - UPLOADS ====================


@app.post("/upload/analise")
async def upload_analise(file: UploadFile = File(...)):
    """Upload do arquivo Analise_Comercial_Completa"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nome de arquivo inválido")

    # Determinar extensão
    ext = Path(file.filename).suffix.lower()
    if ext not in [".xlsx", ".csv"]:
        raise HTTPException(
            status_code=400, detail="Formato inválido. Use .xlsx ou .csv"
        )

    # Salvar na raiz do projeto
    filename = f"Analise_Comercial_Completa{ext}"
    filepath = Path(ROBO_ROOT_PATH) / filename

    async with aiofiles.open(filepath, "wb") as f:
        content = await file.read()
        await f.write(content)

    return {
        "success": True,
        "filename": filename,
        "message": "Arquivo salvo com sucesso",
    }


@app.post("/upload/fin_adcli")
async def upload_fin_adcli(file: UploadFile = File(...)):
    """Upload do arquivo fin_adcli_pg_m3.xls"""
    if not file.filename.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xls")

    filepath = Path(ROBO_ROOT_PATH) / "fin_adcli_pg_m3.xls"
    async with aiofiles.open(filepath, "wb") as f:
        content = await file.read()
        await f.write(content)

    return {"success": True, "filename": "fin_adcli_pg_m3.xls"}


@app.post("/upload/fin_conci")
async def upload_fin_conci(file: UploadFile = File(...)):
    """Upload do arquivo fin_conci_adcli_m3.xls"""
    if not file.filename.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xls")

    filepath = Path(ROBO_ROOT_PATH) / "fin_conci_adcli_m3.xls"
    async with aiofiles.open(filepath, "wb") as f:
        content = await file.read()
        await f.write(content)

    return {"success": True, "filename": "fin_conci_adcli_m3.xls"}


@app.post("/upload/analise_financeira")
async def upload_analise_financeira(file: UploadFile = File(...)):
    """Upload do arquivo Análise Financeira.xlsx"""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xlsx")

    filepath = Path(ROBO_ROOT_PATH) / "Análise Financeira.xlsx"
    async with aiofiles.open(filepath, "wb") as f:
        content = await file.read()
        await f.write(content)

    return {"success": True, "filename": "Análise Financeira.xlsx"}


# ==================== ENDPOINTS - EXECUÇÃO ====================

# Dicionário para armazenar processos ativos
processos_ativos: Dict[str, subprocess.Popen] = {}

# Fases esperadas do cálculo (para estimar progresso)
FASES_CALCULO = [
    ("Iniciando...", 0),
    ("Carregando arquivos...", 10),
    ("Validando dados...", 15),
    ("Pré-processando informações...", 20),
    ("Calculando valores realizados agregados...", 30),
    ("Calculando comissões e FC item a item...", 50),
    ("Carregando estado de recebimentos e aplicando adiantamentos...", 70),
    ("Executando reconciliações de processos quitados...", 85),
    ("Gerando arquivos de saída...", 95),
    ("Concluído", 100),
]


async def monitorar_processo(
    job_id: str, process: subprocess.Popen, mes: int, ano: int
):
    """Monitora processo em background e garante status final."""

    try:
        while process.poll() is None:
            await asyncio.sleep(2)

        return_code = process.returncode

        # Consolidar status final sem sobrescrever o que o processo gerou
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                progress_data = json.load(f)
        except Exception:
            progress_data = {
                "job_id": job_id,
                "percent": 0,
                "etapa": "",
                "mensagens": [],
                "status": "em_andamento",
            }

        if progress_data.get("job_id") != job_id:
            progress_data.update(
                {
                    "job_id": job_id,
                    "percent": 0,
                    "etapa": "",
                    "mensagens": [],
                    "status": "em_andamento",
                }
            )

        if progress_data.get("status") not in ("concluido", "erro"):
            if return_code == 0:
                resultado_path = get_resultado_path()
                etapa_final = "Concluído" if resultado_path else "Processo finalizado"
                progress_data.update(
                    {
                        "etapa": etapa_final,
                        "percent": 100,
                        "status": "concluido",
                    }
                )
            else:
                progress_data.update(
                    {
                        "etapa": f"Processo finalizado (código: {return_code})",
                        "status": "erro",
                        "percent": 100,
                    }
                )

            try:
                with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
                    json.dump(progress_data, f, ensure_ascii=False)
            except Exception:
                pass

    finally:
        processos_ativos.pop(job_id, None)


@app.post("/calcular")
async def iniciar_calculo(
    mes: int = Query(..., ge=1, le=12), ano: int = Query(..., ge=2000, le=2100)
):
    """Inicia cálculo de comissões"""
    job_id = str(uuid.uuid4())

    # Criar arquivo de progresso inicial
    progress_data = {
        "job_id": job_id,
        "percent": 0,
        "etapa": "Iniciando...",
        "mensagens": [],
        "status": "em_andamento",
    }

    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress_data, f, ensure_ascii=False)

    # Disparar subprocesso
    script_path = Path(ROBO_ROOT_PATH) / "calculo_comissoes.py"
    if not script_path.exists():
        raise HTTPException(
            status_code=404, detail="Arquivo calculo_comissoes.py não encontrado"
        )

    # Iniciar processo em background com parâmetros mes/ano
    # Redirecionar stdout/stderr para DEVNULL para evitar bloqueio por buffers cheios
    # O processo não ficará bloqueado esperando que alguém leia os pipes
    env = os.environ.copy()
    env["COMISSOES_JOB_ID"] = job_id
    env["COMISSOES_PROGRESS_FILE"] = PROGRESS_FILE

    process = subprocess.Popen(
        [sys.executable, str(script_path), "--mes", str(mes), "--ano", str(ano)],
        cwd=str(ROBO_ROOT_PATH),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        env=env,
    )

    # Armazenar processo ativo
    processos_ativos[job_id] = process

    # Iniciar monitoramento em background
    asyncio.create_task(monitorar_processo(job_id, process, mes, ano))

    return {"job_id": job_id, "message": "Cálculo iniciado"}


@app.get("/progresso/{job_id}")
async def consultar_progresso(job_id: str):
    """Consulta progresso do cálculo"""
    if not os.path.exists(PROGRESS_FILE):
        # Simular progresso mínimo
        return ProgressResponse(
            job_id=job_id,
            percent=0,
            etapa="Aguardando início...",
            mensagens=[],
            status="em_andamento",
        )

    try:
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            progress = json.load(f)

        # Verificar se job_id corresponde
        if progress.get("job_id") != job_id:
            # Criar novo progresso
            return ProgressResponse(
                job_id=job_id,
                percent=0,
                etapa="Aguardando início...",
                mensagens=[],
                status="em_andamento",
            )

        return ProgressResponse(**progress)
    except Exception as e:
        return ProgressResponse(
            job_id=job_id,
            percent=0,
            etapa=f"Erro ao ler progresso: {str(e)}",
            mensagens=[],
            status="erro",
        )


"""
Novos endpoints para fluxo de Cross-Selling:
 - POST /api/executar-prescan: roda apenas a detecção e retorna casos
 - POST /api/executar-calculo: executa o cálculo completo aceitando decisões
"""


class ExecPrescanRequest(BaseModel):
    mes: int
    ano: int


class ExecCalculoRequest(BaseModel):
    mes: int
    ano: int
    decisoes_cross_selling: Optional[List[Dict[str, Any]]] = None


from contextlib import contextmanager


@contextmanager
def _cwd(path: str):
    old = os.getcwd()
    try:
        os.chdir(path)
        yield
    finally:
        try:
            os.chdir(old)
        except Exception:
            pass


@app.post("/api/executar-prescan")
async def executar_prescan(payload: ExecPrescanRequest):
    import logging
    import traceback

    logger = logging.getLogger(__name__)

    try:
        logger.info(f"[PRESCAN] Iniciando pré-scan para {payload.mes}/{payload.ano}")

        # Instanciar e carregar dados
        from calculo_comissoes import CalculoComissao

        with _cwd(ROBO_ROOT_PATH):
            # Verificar se os arquivos já existem e são recentes (últimos 5 minutos)
            # para evitar reexecutar o preparador desnecessariamente
            # IMPORTANTE: Se o CSV não existir mas o XLSX existir, sempre executar o preparador para converter
            arquivo_csv = Path("Analise_Comercial_Completa.csv")
            arquivo_xlsx = Path("Analise_Comercial_Completa.xlsx")
            precisa_converter = not arquivo_csv.exists() and arquivo_xlsx.exists()

            arquivos_necessarios = [
                Path("Faturados.xlsx"),
                Path("Conversões.xlsx"),
                Path("Faturados_YTD.xlsx"),
            ]
            arquivos_existem = all(f.exists() for f in arquivos_necessarios)
            arquivos_recentes = False
            if arquivos_existem:
                import time

                tempo_atual = time.time()
                arquivos_recentes = all(
                    (tempo_atual - f.stat().st_mtime) < 300  # 5 minutos
                    for f in arquivos_necessarios
                )

            if arquivos_existem and arquivos_recentes and not precisa_converter:
                logger.info(
                    "[PRESCAN] Arquivos já existem e são recentes, pulando preparador"
                )
            else:
                if precisa_converter:
                    logger.info(
                        "[PRESCAN] Arquivo XLSX detectado sem CSV correspondente, executando preparador para conversão"
                    )
                logger.info("[PRESCAN] Executando preparador de dados...")
                # Preparar dados do mês/ano antes do pré-scan (necessário para popular FATURADOS etc.)
                try:
                    import preparar_dados_mensais

                    preparar_dados_mensais.run_preparador(payload.mes, payload.ano)
                    logger.info("[PRESCAN] Preparador concluído")
                except Exception as e:
                    logger.error(
                        f"[PRESCAN] Erro no preparador: {e}\n{traceback.format_exc()}"
                    )
                    # Não abortar, mas registrar no detalhe do erro do endpoint se falhar
                    raise HTTPException(
                        status_code=500, detail=f"Falha no preparador: {e}"
                    )

            logger.info("[PRESCAN] Configurando caminhos de arquivos...")
            # Ajustar caminhos como no CLI principal
            try:
                import calculo_comissoes as cc

                cc.ARQUIVO_FATURADOS = "Faturados.xlsx"
                cc.ARQUIVO_CONVERSOES = "Conversões.xlsx"
                cc.ARQUIVO_FATURADOS_YTD = "Faturados_YTD.xlsx"
                mm = str(payload.mes).zfill(2)
                import glob as _glob

                encontrados = _glob.glob(
                    str(Path("rentabilidades") / f"*{mm}*{payload.ano}*agrupada*.xlsx")
                )
                if encontrados:
                    cc.ARQUIVO_RENTABILIDADE = encontrados[0]
                    logger.info(
                        f"[PRESCAN] Arquivo de rentabilidade: {cc.ARQUIVO_RENTABILIDADE}"
                    )
                else:
                    padrao = (
                        Path("rentabilidades")
                        / f"rentabilidade_{mm}_{payload.ano}_agrupada.xlsx"
                    )
                    if padrao.exists():
                        cc.ARQUIVO_RENTABILIDADE = str(padrao)
                        logger.info(
                            f"[PRESCAN] Arquivo de rentabilidade: {cc.ARQUIVO_RENTABILIDADE}"
                        )
                    else:
                        # Se não encontrou, deixar None (o código defensivo tratará)
                        cc.ARQUIVO_RENTABILIDADE = None
                        logger.warning(
                            f"[PRESCAN] Arquivo de rentabilidade não encontrado para {mm}/{payload.ano}"
                        )
            except Exception as e:
                logger.error(
                    f"[PRESCAN] Erro ao configurar caminhos: {e}\n{traceback.format_exc()}"
                )
                # Se houver erro, garantir que None seja definido para evitar NameError
                cc.ARQUIVO_RENTABILIDADE = None

            logger.info("[PRESCAN] Instanciando CalculoComissao...")
            calc = CalculoComissao()

            logger.info("[PRESCAN] Carregando dados...")
            # Carregar e pré-processar dados existentes
            calc._carregar_dados()
            logger.info("[PRESCAN] Pré-processando dados...")
            calc._preprocessar_dados()

            logger.info("[PRESCAN] Rodando detecção de cross-selling...")
            # Rodar apenas detecção
            calc._detectar_cross_selling()

        casos = getattr(calc, "casos_cross_selling_detectados", []) or []
        logger.info(f"[PRESCAN] Detecção concluída: {len(casos)} caso(s) encontrado(s)")

        # Normalizar tipos simples
        out = []
        for c in casos:
            out.append(
                {
                    "processo": str(c.get("processo")),
                    "consultor": c.get("consultor"),
                    "linha": c.get("linha"),
                    "taxa": float(c.get("taxa", 0.0)),
                }
            )
        logger.info(f"[PRESCAN] Retornando {len(out)} caso(s)")
        return out
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[PRESCAN] Erro fatal: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Erro no pré-scan: {str(e)}")


@app.post("/api/executar-calculo")
async def executar_calculo(payload: ExecCalculoRequest):
    try:
        # Execução síncrona com decisões vindas da UI
        from calculo_comissoes import CalculoComissao

        with _cwd(ROBO_ROOT_PATH):
            # Preparar dados do mês/ano antes da execução
            try:
                import preparar_dados_mensais

                preparar_dados_mensais.run_preparador(payload.mes, payload.ano)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Falha no preparador: {e}")

            # Ajustar caminhos como no CLI principal
            try:
                import calculo_comissoes as cc

                cc.ARQUIVO_FATURADOS = "Faturados.xlsx"
                cc.ARQUIVO_CONVERSOES = "Conversões.xlsx"
                cc.ARQUIVO_FATURADOS_YTD = "Faturados_YTD.xlsx"
                mm = str(payload.mes).zfill(2)
                import glob as _glob

                encontrados = _glob.glob(
                    str(Path("rentabilidades") / f"*{mm}*{payload.ano}*agrupada*.xlsx")
                )
                if encontrados:
                    cc.ARQUIVO_RENTABILIDADE = encontrados[0]
                else:
                    padrao = (
                        Path("rentabilidades")
                        / f"rentabilidade_{mm}_{payload.ano}_agrupada.xlsx"
                    )
                    if padrao.exists():
                        cc.ARQUIVO_RENTABILIDADE = str(padrao)
                    else:
                        # Se não encontrou, deixar None (o código defensivo tratará)
                        cc.ARQUIVO_RENTABILIDADE = None
            except Exception:
                # Se houver erro, garantir que None seja definido para evitar NameError
                cc.ARQUIVO_RENTABILIDADE = None

            calc = CalculoComissao()
            calc.executar(decisoes_cross_selling=payload.decisoes_cross_selling or [])
        return {"success": True, "message": "Cálculo concluído"}
    except HTTPException:
        raise
    except Exception as e:
        import traceback

        erro_completo = traceback.format_exc()
        print(f"[adapter] ERRO ao executar cálculo:\n{erro_completo}")
        raise HTTPException(
            status_code=500, detail=f"Erro ao executar cálculo: {str(e)}"
        )


# ==================== ENDPOINTS - DEBUG ====================


@app.get("/debug/logs")
async def obter_logs(lines: int = Query(200, ge=1, le=5000)):
    """Retorna as últimas linhas do arquivo de logs do adapter."""
    try:
        log_path = Path(LOG_FILE)
        if not log_path.exists():
            return PlainTextResponse("Arquivo de log não encontrado.", status_code=404)
        with open(log_path, "r", encoding="utf-8") as f:
            conteudo = f.readlines()
        tail = "".join(conteudo[-lines:])
        return PlainTextResponse(tail)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler logs: {e}")


# ==================== ENDPOINTS - RESULTADOS ====================


@app.get("/resultado/abas")
async def listar_abas_resultado():
    """Lista abas do arquivo de resultado mais recente"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        print("[adapter] /resultado/abas -> nenhum arquivo de resultado encontrado")
        return {"abas": []}

    try:
        wb = load_workbook(resultado_path, read_only=True)
        abas = wb.sheetnames
        print(f"[adapter] /resultado/abas -> arquivo={resultado_path.name} abas={abas}")
        return {"abas": abas, "arquivo": resultado_path.name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao ler resultado: {str(e)}")


@app.get("/resultado/aba/{nome_aba}")
async def ler_aba_resultado(
    nome_aba: str,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=1000),
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = Query("asc", regex="^(asc|desc)$"),
    filters: Optional[str] = None,
):
    """Lê uma aba do resultado com paginação"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        raise HTTPException(
            status_code=404, detail="Nenhum arquivo de resultado encontrado"
        )

    print(f"[adapter] /resultado/aba/{nome_aba} -> arquivo={resultado_path.name}")
    df = read_excel_sheet(resultado_path, nome_aba)
    print(
        f"[adapter] /resultado/aba/{nome_aba} -> linhas={len(df)} colunas={list(df.columns)}"
    )

    # Aplicar filtros
    if filters:
        try:
            filter_dict = json.loads(filters)
            for col, value in filter_dict.items():
                if col in df.columns and value:
                    df = df[
                        df[col]
                        .astype(str)
                        .str.contains(str(value), case=False, na=False)
                    ]
        except Exception:
            pass

    # Ordenação
    if sort_by and sort_by in df.columns:
        ascending = sort_order == "asc"
        df = df.sort_values(by=sort_by, ascending=ascending)

    # Paginação
    total = len(df)
    start = (page - 1) * size
    end = start + size
    df_page = df.iloc[start:end]

    return {
        "data": df_page.to_dict(orient="records"),
        "total": total,
        "page": page,
        "size": size,
        "columns": list(df.columns),
        "arquivo": resultado_path.name,
    }


@app.get("/resultado/aba/{nome_aba}/valores-unicos/{coluna}")
async def obter_valores_unicos_resultado(nome_aba: str, coluna: str):
    """Retorna valores únicos de uma coluna específica da aba de resultado"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        raise HTTPException(
            status_code=404, detail="Nenhum arquivo de resultado encontrado"
        )

    df = read_excel_sheet(resultado_path, nome_aba)

    if coluna not in df.columns:
        raise HTTPException(
            status_code=404,
            detail=f"Coluna '{coluna}' não encontrada na aba '{nome_aba}'",
        )

    # Extrair valores únicos, remover NaN e valores vazios, converter para string e ordenar
    valores_unicos = (
        df[coluna]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", None)
        .dropna()
        .unique()
        .tolist()
    )
    valores_unicos.sort()

    return {"coluna": coluna, "valores": valores_unicos}


@app.get("/baixar/resultado")
async def baixar_resultado():
    """Download do arquivo de resultado mais recente"""
    resultado_path = get_resultado_path()
    if not resultado_path:
        raise HTTPException(
            status_code=404, detail="Nenhum arquivo de resultado encontrado"
        )

    return FileResponse(
        resultado_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=resultado_path.name,
    )


@app.get("/health")
async def health_check():
    """Health check"""
    regras_path = get_regras_path()
    return {
        "status": "ok",
        "robo_path": ROBO_ROOT_PATH,
        "regras_path": str(regras_path),
        "regras_exists": regras_path.exists(),
        "resultado_exists": get_resultado_path() is not None,
        "env_file_exists": (adapter_dir / ".env").exists(),
    }


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
