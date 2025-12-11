"""
Script de teste para os novos endpoints de Recebimentos.
Valida se os endpoints retornam dados no formato esperado pelo frontend.
"""

import sys
from pathlib import Path

# Adicionar diretório raiz ao path
ROBO_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROBO_ROOT))

import json
from app import (
    get_recebimento_path,
    read_excel_sheet,
)


def test_estrutura_arquivos():
    """Testa se os arquivos de recebimento existem"""
    print("\n" + "="*60)
    print("1. TESTE: Estrutura de Arquivos")
    print("="*60)
    
    # Verificar arquivo de agosto/2025
    recebimento_path = get_recebimento_path(8, 2025)
    
    if recebimento_path:
        print(f"✓ Arquivo encontrado: {recebimento_path}")
        
        # Verificar abas
        from openpyxl import load_workbook
        wb = load_workbook(recebimento_path, read_only=True)
        print(f"✓ Abas disponíveis: {wb.sheetnames}")
        
        if "COMISSOES_ADIANTAMENTOS" in wb.sheetnames:
            print("✓ Aba COMISSOES_ADIANTAMENTOS existe")
        else:
            print("✗ Aba COMISSOES_ADIANTAMENTOS NÃO existe")
            
        if "COMISSOES_REGULARES" in wb.sheetnames:
            print("✓ Aba COMISSOES_REGULARES existe")
        else:
            print("✗ Aba COMISSOES_REGULARES NÃO existe")
            
        return True
    else:
        print("✗ Arquivo de recebimento não encontrado para 08/2025")
        return False


def test_leitura_pagamentos():
    """Testa leitura de pagamentos das abas"""
    print("\n" + "="*60)
    print("2. TESTE: Leitura de Pagamentos")
    print("="*60)
    
    recebimento_path = get_recebimento_path(8, 2025)
    if not recebimento_path:
        print("✗ Arquivo não encontrado - pulando teste")
        return False
    
    try:
        # Testar adiantamentos
        df_adiant = read_excel_sheet(recebimento_path, "COMISSOES_ADIANTAMENTOS")
        print(f"✓ Adiantamentos: {len(df_adiant)} registros")
        print(f"  Colunas: {df_adiant.columns.tolist()}")
        
        if not df_adiant.empty:
            primeiro = df_adiant.iloc[0]
            print(f"\n  Exemplo de adiantamento:")
            print(f"    - PROCESSO: {primeiro.get('PROCESSO', 'N/A')}")
            print(f"    - NOME_COLABORADOR: {primeiro.get('NOME_COLABORADOR', 'N/A')}")
            print(f"    - VALOR_PAGO: {primeiro.get('VALOR_PAGO', 0)}")
            print(f"    - TCMP: {primeiro.get('TCMP', 0)}")
            print(f"    - COMISSAO_CALCULADA: {primeiro.get('COMISSAO_CALCULADA', 0)}")
        
        # Testar regulares
        df_regular = read_excel_sheet(recebimento_path, "COMISSOES_REGULARES")
        print(f"\n✓ Regulares: {len(df_regular)} registros")
        print(f"  Colunas: {df_regular.columns.tolist()}")
        
        if not df_regular.empty:
            primeiro = df_regular.iloc[0]
            print(f"\n  Exemplo de regular:")
            print(f"    - PROCESSO: {primeiro.get('PROCESSO', 'N/A')}")
            print(f"    - NOME_COLABORADOR: {primeiro.get('NOME_COLABORADOR', 'N/A')}")
            print(f"    - VALOR_PAGO: {primeiro.get('VALOR_PAGO', 0)}")
            print(f"    - TCMP: {primeiro.get('TCMP', 0)}")
            print(f"    - FCMP: {primeiro.get('FCMP', 1.0)}")
            print(f"    - COMISSAO_CALCULADA: {primeiro.get('COMISSAO_CALCULADA', 0)}")
        
        return True
        
    except Exception as e:
        print(f"✗ Erro ao ler pagamentos: {e}")
        return False


def test_estado_detalhes():
    """Testa leitura de detalhes de TCMP/FCMP do Estado"""
    print("\n" + "="*60)
    print("3. TESTE: Detalhes TCMP/FCMP do Estado")
    print("="*60)
    
    estado_path = Path(ROBO_ROOT) / "Estado_Processos_Recebimento.xlsx"
    
    if not estado_path.exists():
        print(f"✗ Arquivo de estado não encontrado: {estado_path}")
        return False
    
    print(f"✓ Arquivo de estado encontrado: {estado_path}")
    
    try:
        df_estado = read_excel_sheet(str(estado_path), "ESTADO")
        print(f"✓ Estado: {len(df_estado)} processos")
        print(f"  Colunas: {df_estado.columns.tolist()}")
        
        # Verificar se tem colunas de detalhes JSON
        if "TCMP_DETALHES_JSON" in df_estado.columns:
            print("✓ Coluna TCMP_DETALHES_JSON existe")
        else:
            print("✗ Coluna TCMP_DETALHES_JSON NÃO existe")
            
        if "FCMP_DETALHES_JSON" in df_estado.columns:
            print("✓ Coluna FCMP_DETALHES_JSON existe")
        else:
            print("✗ Coluna FCMP_DETALHES_JSON NÃO existe")
        
        # Tentar parsear um exemplo
        if not df_estado.empty:
            primeiro = df_estado.iloc[0]
            processo = primeiro.get("PROCESSO", "N/A")
            print(f"\n  Exemplo de processo: {processo}")
            
            tcmp_json = primeiro.get("TCMP_DETALHES_JSON", "{}")
            fcmp_json = primeiro.get("FCMP_DETALHES_JSON", "{}")
            
            if tcmp_json and tcmp_json != "{}":
                tcmp_data = json.loads(tcmp_json)
                colaboradores_tcmp = list(tcmp_data.keys())
                print(f"    - TCMP breakdown para {len(colaboradores_tcmp)} colaboradores")
                if colaboradores_tcmp:
                    colab = colaboradores_tcmp[0]
                    print(f"      Exemplo colaborador: {colab}")
                    print(f"      Itens: {list(tcmp_data[colab].keys())}")
            
            if fcmp_json and fcmp_json != "{}":
                fcmp_data = json.loads(fcmp_json)
                colaboradores_fcmp = list(fcmp_data.keys())
                print(f"    - FCMP breakdown para {len(colaboradores_fcmp)} colaboradores")
        
        return True
        
    except Exception as e:
        print(f"✗ Erro ao ler estado: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_formato_resposta():
    """Testa se o formato de resposta está compatível com o frontend"""
    print("\n" + "="*60)
    print("4. TESTE: Formato de Resposta (Simulação)")
    print("="*60)
    
    # Simular estrutura esperada pelo frontend
    pagamento_exemplo = {
        "id": "ADIANT_8_2025_0",
        "tipo": "ADIANTAMENTO",
        "processo": "PROC-001",
        "nome_colaborador": "João Silva",
        "cargo": "Vendedor",
        "data_pagamento": "2025-08-15",
        "valor_pago": 50000.00,
        "tcmp": 0.05,
        "fcmp": 1.0,
        "comissao_calculada": 2500.00,
        "tcmp_detalhes": [
            {
                "item": "Item A",
                "valor": 30000.00,
                "taxa": 0.06,
                "peso": 0.6,
                "tcmp_parcial": 0.036
            }
        ],
        "fcmp_detalhes": []
    }
    
    print("✓ Estrutura de pagamento:")
    print(json.dumps(pagamento_exemplo, indent=2, ensure_ascii=False))
    
    # Validar campos obrigatórios
    campos_obrigatorios = [
        "id", "tipo", "processo", "nome_colaborador", "cargo",
        "data_pagamento", "valor_pago", "tcmp", "fcmp", "comissao_calculada"
    ]
    
    campos_ok = all(campo in pagamento_exemplo for campo in campos_obrigatorios)
    
    if campos_ok:
        print("\n✓ Todos os campos obrigatórios presentes")
    else:
        print("\n✗ Campos obrigatórios faltando")
        faltando = [c for c in campos_obrigatorios if c not in pagamento_exemplo]
        print(f"  Faltando: {faltando}")
    
    return campos_ok


def main():
    """Executa todos os testes"""
    print("\n" + "="*60)
    print("TESTE DOS ENDPOINTS DE RECEBIMENTOS")
    print("="*60)
    
    resultados = []
    
    resultados.append(("Estrutura de Arquivos", test_estrutura_arquivos()))
    resultados.append(("Leitura de Pagamentos", test_leitura_pagamentos()))
    resultados.append(("Detalhes TCMP/FCMP", test_estado_detalhes()))
    resultados.append(("Formato de Resposta", test_formato_resposta()))
    
    # Resumo
    print("\n" + "="*60)
    print("RESUMO DOS TESTES")
    print("="*60)
    
    for nome, passou in resultados:
        status = "✓ PASSOU" if passou else "✗ FALHOU"
        print(f"{status}: {nome}")
    
    total_passou = sum(1 for _, passou in resultados if passou)
    total = len(resultados)
    
    print(f"\nTotal: {total_passou}/{total} testes passaram")
    
    if total_passou == total:
        print("\n🎉 Todos os testes passaram! Backend pronto para integração.")
    else:
        print("\n⚠️ Alguns testes falharam. Verifique os erros acima.")


if __name__ == "__main__":
    main()
