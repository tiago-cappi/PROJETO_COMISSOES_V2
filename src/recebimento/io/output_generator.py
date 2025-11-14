"""
Gera arquivo Excel de saída para comissões por recebimento.
"""

import pandas as pd
import os
from typing import Dict, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class RecebimentoOutputGenerator:
    """
    Gera arquivo Excel com todas as abas de comissões por recebimento.
    """
    
    def __init__(self):
        """Inicializa o gerador de saída."""
        pass
    
    def gerar(
        self,
        mes: int,
        ano: int,
        dados: Dict[str, pd.DataFrame],
        base_path: str = "."
    ) -> str:
        """
        Gera arquivo Excel com todas as abas.
        
        Args:
            mes: Mês de apuração (1-12)
            ano: Ano de apuração (ex: 2025)
            dados: Dict com DataFrames:
                - 'adiantamentos': DataFrame com comissões de adiantamentos
                - 'regulares': DataFrame com comissões regulares
                - 'reconciliacoes': DataFrame com reconciliações (pode estar vazio)
                - 'estado': DataFrame com estado completo
                - 'avisos': DataFrame com documentos não mapeados
            base_path: Caminho base para salvar o arquivo
        
        Returns:
            Caminho do arquivo gerado
        """
        filename = f"Comissoes_Recebimento_{mes:02d}_{ano}.xlsx"
        filepath = os.path.join(base_path, filename)
        
        print(f"[RECEBIMENTO] [OUTPUT] Iniciando geração do arquivo Excel...")
        print(f"[RECEBIMENTO] [OUTPUT] Nome do arquivo: {filename}")
        print(f"[RECEBIMENTO] [OUTPUT] Caminho completo: {filepath}")
        print(f"[RECEBIMENTO] [OUTPUT] Base path: {base_path}")
        
        # Criar ExcelWriter
        print(f"[RECEBIMENTO] [OUTPUT] Criando ExcelWriter...")
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Aba 1: COMISSOES_ADIANTAMENTOS
                print(f"[RECEBIMENTO] [OUTPUT] Criando aba COMISSOES_ADIANTAMENTOS...")
                if not dados.get('adiantamentos', pd.DataFrame()).empty:
                    df_adiant = dados['adiantamentos'].copy()
                    print(f"[RECEBIMENTO] [OUTPUT] Adiantamentos: {len(df_adiant)} linha(s)")
                    self._preparar_dataframe_adiantamentos(df_adiant)
                    df_adiant.to_excel(
                        writer,
                        sheet_name='COMISSOES_ADIANTAMENTOS',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba COMISSOES_ADIANTAMENTOS criada com sucesso")
                else:
                    print(f"[RECEBIMENTO] [OUTPUT] Nenhum adiantamento. Criando aba vazia...")
                    # Criar DataFrame vazio com colunas esperadas
                    df_vazio = pd.DataFrame(columns=[
                        'processo', 'documento', 'data_pagamento', 'valor_pago',
                        'nome_colaborador', 'cargo', 'tcmp', 'fc', 'comissao_calculada',
                        'mes_calculo', 'observacao'
                    ])
                    df_vazio.to_excel(
                        writer,
                        sheet_name='COMISSOES_ADIANTAMENTOS',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba COMISSOES_ADIANTAMENTOS vazia criada")
                
                # Aba 2: COMISSOES_REGULARES
                print(f"[RECEBIMENTO] [OUTPUT] Criando aba COMISSOES_REGULARES...")
                if not dados.get('regulares', pd.DataFrame()).empty:
                    df_reg = dados['regulares'].copy()
                    print(f"[RECEBIMENTO] [OUTPUT] Regulares: {len(df_reg)} linha(s)")
                    self._preparar_dataframe_regulares(df_reg)
                    df_reg.to_excel(
                        writer,
                        sheet_name='COMISSOES_REGULARES',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba COMISSOES_REGULARES criada com sucesso")
                else:
                    print(f"[RECEBIMENTO] [OUTPUT] Nenhum pagamento regular. Criando aba vazia...")
                    # Criar DataFrame vazio com colunas esperadas
                    df_vazio = pd.DataFrame(columns=[
                        'processo', 'documento', 'data_pagamento', 'valor_pago',
                        'nome_colaborador', 'cargo', 'tcmp', 'fcmp', 'comissao_calculada',
                        'mes_faturamento', 'mes_calculo', 'observacao'
                    ])
                    df_vazio.to_excel(
                        writer,
                        sheet_name='COMISSOES_REGULARES',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba COMISSOES_REGULARES vazia criada")
                
                # Aba 3: RECONCILIACOES
                print(f"[RECEBIMENTO] [OUTPUT] Criando aba RECONCILIACOES...")
                if not dados.get('reconciliacoes', pd.DataFrame()).empty:
                    dados['reconciliacoes'].to_excel(
                        writer,
                        sheet_name='RECONCILIACOES',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba RECONCILIACOES criada")
                else:
                    print(f"[RECEBIMENTO] [OUTPUT] Nenhuma reconciliação. Criando aba vazia...")
                    # Criar DataFrame vazio com colunas esperadas
                    df_vazio = pd.DataFrame(columns=[
                        'processo',
                        'colaborador',
                        'tcmp',
                        'fcmp',
                        'comissao_adiantada_fc_1',
                        'comissao_deveria_fc_real',
                        'diferenca_fc',
                        'ajuste_reconciliacao',
                        'mes_faturamento',
                    ])
                    df_vazio.to_excel(
                        writer,
                        sheet_name='RECONCILIACOES',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba RECONCILIACOES vazia criada")
                
                # Aba 4: ESTADO
                print(f"[RECEBIMENTO] [OUTPUT] Criando aba ESTADO...")
                if not dados.get('estado', pd.DataFrame()).empty:
                    print(f"[RECEBIMENTO] [OUTPUT] Estado: {len(dados['estado'])} linha(s)")
                    dados['estado'].to_excel(
                        writer,
                        sheet_name='ESTADO',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba ESTADO criada com sucesso")
                else:
                    print(f"[RECEBIMENTO] [OUTPUT] Estado vazio. Criando aba vazia...")
                    # Criar DataFrame vazio com colunas do schema
                    from ..estado.state_schema import COLUNAS_ESTADO
                    df_vazio = pd.DataFrame(columns=COLUNAS_ESTADO)
                    df_vazio.to_excel(
                        writer,
                        sheet_name='ESTADO',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba ESTADO vazia criada")
                
                # Aba 5: AVISOS
                print(f"[RECEBIMENTO] [OUTPUT] Criando aba AVISOS...")
                if not dados.get('avisos', pd.DataFrame()).empty:
                    print(f"[RECEBIMENTO] [OUTPUT] Avisos: {len(dados['avisos'])} linha(s)")
                    dados['avisos'].to_excel(
                        writer,
                        sheet_name='AVISOS',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba AVISOS criada com sucesso")
                else:
                    print(f"[RECEBIMENTO] [OUTPUT] Nenhum aviso. Criando aba vazia...")
                    # Criar DataFrame vazio
                    df_vazio = pd.DataFrame(columns=[
                        'documento', 'documento_6dig', 'motivo', 'valor', 'data_pagamento'
                    ])
                    df_vazio.to_excel(
                        writer,
                        sheet_name='AVISOS',
                        index=False
                    )
                    print(f"[RECEBIMENTO] [OUTPUT] Aba AVISOS vazia criada")
        
        except Exception as e:
            print(f"[RECEBIMENTO] [OUTPUT] ERRO ao gerar arquivo Excel: {e}")
            import traceback
            traceback.print_exc()
            raise
        
        print(f"[RECEBIMENTO] [OUTPUT] ExcelWriter fechado. Arquivo salvo.")
        
        # Aplicar formatação
        print(f"[RECEBIMENTO] [OUTPUT] Aplicando formatação...")
        self._aplicar_formatacao(filepath)
        print(f"[RECEBIMENTO] [OUTPUT] Formatação aplicada")
        
        print(f"[RECEBIMENTO] [OUTPUT] Arquivo gerado com sucesso: {filepath}")
        return filepath
    
    def _preparar_dataframe_adiantamentos(self, df: pd.DataFrame):
        """Prepara DataFrame de adiantamentos para exibição."""
        # Ordenar por processo e colaborador
        if not df.empty:
            if 'processo' in df.columns and 'nome_colaborador' in df.columns:
                df.sort_values(['processo', 'nome_colaborador'], inplace=True)
            
            # Formatar valores monetários
            if 'valor_pago' in df.columns:
                df['valor_pago'] = df['valor_pago'].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "")
            if 'comissao_calculada' in df.columns:
                df['comissao_calculada'] = df['comissao_calculada'].apply(
                    lambda x: f"{x:,.2f}" if pd.notna(x) else ""
                )
            
            # Formatar datas
            if 'data_pagamento' in df.columns:
                df['data_pagamento'] = pd.to_datetime(df['data_pagamento'], errors='coerce')
                df['data_pagamento'] = df['data_pagamento'].dt.strftime('%d/%m/%Y')
    
    def _preparar_dataframe_regulares(self, df: pd.DataFrame):
        """Prepara DataFrame de pagamentos regulares para exibição."""
        # Ordenar por processo e colaborador
        if not df.empty:
            if 'processo' in df.columns and 'nome_colaborador' in df.columns:
                df.sort_values(['processo', 'nome_colaborador'], inplace=True)
            
            # Formatar valores monetários
            if 'valor_pago' in df.columns:
                df['valor_pago'] = df['valor_pago'].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "")
            if 'comissao_calculada' in df.columns:
                df['comissao_calculada'] = df['comissao_calculada'].apply(
                    lambda x: f"{x:,.2f}" if pd.notna(x) else ""
                )
            
            # Formatar datas
            if 'data_pagamento' in df.columns:
                df['data_pagamento'] = pd.to_datetime(df['data_pagamento'], errors='coerce')
                df['data_pagamento'] = df['data_pagamento'].dt.strftime('%d/%m/%Y')
    
    def _aplicar_formatacao(self, filepath: str):
        """
        Aplica formatação básica ao arquivo Excel.
        
        Args:
            filepath: Caminho do arquivo Excel
        """
        try:
            wb = load_workbook(filepath)

            # Cores para destacar tipos
            fill_adiantamento = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            fill_regular = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
            fill_reconciliacao = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            fill_header = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
            font_header = Font(bold=True, color="FFFFFF")
            
            # Formatar cada aba
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Formatar cabeçalho
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.fill = fill_header
                        cell.font = font_header
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Ajustar largura das colunas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Aplicar cores condicionais por tipo de lançamento
                    if sheet_name == "COMISSOES_ADIANTAMENTOS":
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for cell in row:
                                cell.fill = fill_adiantamento
                    elif sheet_name == "COMISSOES_REGULARES":
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for cell in row:
                                cell.fill = fill_regular
                    elif sheet_name == "RECONCILIACOES":
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for cell in row:
                                cell.fill = fill_reconciliacao
            
            wb.save(filepath)
        except Exception:
            # Se houver erro na formatação, arquivo ainda será salvo sem formatação
            pass

