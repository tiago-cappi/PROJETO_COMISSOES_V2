"""
Módulo de estilização de planilhas Excel.
Contém funções para aplicar cores e formatação nas abas de saída.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def light_fill(rgb_hex: str) -> PatternFill:
    """
    Cria um PatternFill com cor clara (pastel).
    
    Args:
        rgb_hex: Código hexadecimal da cor sem '#' (ex: 'E3F2FD')
    
    Returns:
        PatternFill configurado com a cor especificada.
    """
    return PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")


# Paleta suave (pastéis, claras) para coloração de grupos de colunas
PALETTE = [
    "E3F2FD",  # azul claríssimo
    "E8F5E9",  # verde claríssimo
    "FFF8E1",  # amarelo claríssimo
    "F3E5F5",  # lilás claríssimo
    "E0F7FA",  # ciano claríssimo
    "FBE9E7",  # pêssego claríssimo
    "F1F8E9",  # lima claríssimo
    "FFF3E0",  # laranja claríssimo
    "EDE7F6",  # roxo claríssimo
    "F9FBE7",  # chartreuse claríssimo
]


# Grupo → padrões de identificação de coluna (qualquer substring no cabeçalho)
FC_GROUP_PATTERNS = {
    # faturamento (linha / individual)
    "faturamento_linha": [
        "FATURAMENTO_LINHA",
        "FC_FATURAMENTO_LINHA",
        "PESO_FATURAMENTO_LINHA",
        "META_FATURAMENTO_LINHA",
        "REALIZADO_FATURAMENTO_LINHA",
        "ATINGIMENTO_FATURAMENTO_LINHA",
        "PESO_FAT_LINHA",
        "META_FAT_LINHA",
        "REALIZADO_FAT_LINHA",
        "ATING_FAT_LINHA",
        "ATING_CAP_FAT_LINHA",
        "COMP_FC_FAT_LINHA",
    ],
    "faturamento_individual": [
        "FATURAMENTO_INDIVIDUAL",
        "FC_FATURAMENTO_INDIVIDUAL",
        "PESO_FATURAMENTO_INDIVIDUAL",
        "META_FATURAMENTO_INDIVIDUAL",
        "REALIZADO_FATURAMENTO_INDIVIDUAL",
        "ATINGIMENTO_FATURAMENTO_INDIVIDUAL",
        "PESO_FAT_IND",
        "META_FAT_IND",
        "REALIZADO_FAT_IND",
        "ATING_FAT_IND",
        "ATING_CAP_FAT_IND",
        "COMP_FC_FAT_IND",
    ],
    # conversão (linha / individual)
    "conversao_linha": [
        "CONVERSAO_LINHA",
        "FC_CONVERSAO_LINHA",
        "PESO_CONVERSAO_LINHA",
        "META_CONVERSAO_LINHA",
        "REALIZADO_CONVERSAO_LINHA",
        "ATINGIMENTO_CONVERSAO_LINHA",
        "PESO_CONV_LINHA",
        "META_CONV_LINHA",
        "REALIZADO_CONV_LINHA",
        "ATING_CONV_LINHA",
        "ATING_CAP_CONV_LINHA",
        "COMP_FC_CONV_LINHA",
    ],
    "conversao_individual": [
        "CONVERSAO_INDIVIDUAL",
        "FC_CONVERSAO_INDIVIDUAL",
        "PESO_CONVERSAO_INDIVIDUAL",
        "META_CONVERSAO_INDIVIDUAL",
        "REALIZADO_CONVERSAO_INDIVIDUAL",
        "ATINGIMENTO_CONVERSAO_INDIVIDUAL",
        "PESO_CONV_IND",
        "META_CONV_IND",
        "REALIZADO_CONV_IND",
        "ATING_CONV_IND",
        "ATING_CAP_CONV_IND",
        "COMP_FC_CONV_IND",
    ],
    # rentabilidade / retenção
    "rentabilidade": [
        "RENTABILIDADE",
        "FC_RENTABILIDADE",
        "PESO_RENTABILIDADE",
        "META_RENTABILIDADE",
        "REALIZADO_RENTABILIDADE",
        "ATINGIMENTO_RENTABILIDADE",
        "PESO_RENTAB",
        "META_RENTAB",
        "REALIZADO_RENTAB",
        "ATING_RENTAB",
        "ATING_CAP_RENTAB",
        "COMP_FC_RENTAB",
    ],
    "retencao": [
        "RETENCAO",
        "FC_RETENCAO",
        "PESO_RETENCAO",
        "META_RETENCAO",
        "REALIZADO_RETENCAO",
        "ATINGIMENTO_RETENCAO",
        "PESO_RETENCAO",
        "META_RETENCAO",
        "REALIZADO_RETENCAO",
        "ATING_RETENCAO",
        "ATING_CAP_RETENCAO",
        "COMP_FC_RETENCAO",
    ],
    # metas de fornecedores (se existirem)
    "fornecedor_1": [
        "FORNECEDOR_1",
        "FC_FORNECEDOR_1",
        "PESO_FORNECEDOR_1",
        "META_FORNECEDOR_1",
        "REALIZADO_FORNECEDOR_1",
        "ATINGIMENTO_FORNECEDOR_1",
        "PESO_FORN1",
        "META_FORN1",
        "REALIZADO_FORN1",
        "ATING_FORN1",
        "ATING_CAP_FORN1",
        "COMP_FC_FORN1",
        "MOEDA_FORN1",
    ],
    "fornecedor_2": [
        "FORNECEDOR_2",
        "FC_FORNECEDOR_2",
        "PESO_FORNECEDOR_2",
        "META_FORNECEDOR_2",
        "REALIZADO_FORNECEDOR_2",
        "ATINGIMENTO_FORNECEDOR_2",
        "PESO_FORN2",
        "META_FORN2",
        "REALIZADO_FORN2",
        "ATING_FORN2",
        "ATING_CAP_FORN2",
        "COMP_FC_FORN2",
        "MOEDA_FORN2",
    ],
    # FC total (às vezes reportado em colunas próprias)
    "fc_total": [
        "FC_TOTAL",
        "FC_GERAL",
        "FATOR_CORRECAO_FC",
        "FATOR_CORRECAO",
        "FC",
    ],
}


def match_group(header: str):
    """
    Identifica a qual grupo de padrões um cabeçalho de coluna pertence.
    
    Args:
        header: Nome do cabeçalho da coluna
    
    Returns:
        Nome do grupo encontrado ou None se não houver correspondência.
    """
    h = header.upper().strip()
    for group, patterns in FC_GROUP_PATTERNS.items():
        for p in patterns:
            if p in h:
                return group
    return None


def apply_group_fills_to_sheet(ws):
    """
    Pinta colunas inteiras do mesmo grupo (componente FC/meta/realizado/peso/atingimento)
    com a mesma cor clara. Não altera valores.
    
    Args:
        ws: Worksheet do openpyxl a ser estilizada
    """
    header_row = 1
    col_group = {}
    max_col = ws.max_column
    max_row = ws.max_row

    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        group = match_group(str(cell.value) if cell.value is not None else "")
        if group:
            col_group[c] = group

    if not col_group:
        return

    group_keys = sorted(set(col_group.values()))
    color_for_group = {}
    for idx, g in enumerate(group_keys):
        color_for_group[g] = light_fill(PALETTE[idx % len(PALETTE)])

    for c, g in col_group.items():
        fill = color_for_group[g]
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=c).fill = fill


def style_output_workbook(xlsx_path: str):
    """
    Carrega o arquivo Excel gerado e aplica a coloração de grupos
    nas abas COMISSOES_CALCULADAS e RECONCILIACOES/RECONCILIACAO.
    
    Args:
        xlsx_path: Caminho completo para o arquivo Excel a ser estilizado
    """
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        return

    targets = []
    for name in ("COMISSOES_CALCULADAS", "RECONCILIACOES", "RECONCILIACAO"):
        if name in wb.sheetnames:
            targets.append(name)

    for sheet_name in targets:
        ws = wb[sheet_name]
        apply_group_fills_to_sheet(ws)

    wb.save(xlsx_path)

