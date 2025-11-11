"""
Arquivo de testes temporário para validar as funções utilitárias migradas.
Execute este arquivo para verificar se todas as funções estão funcionando corretamente.
"""

import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Adicionar o diretório raiz ao path para importar os módulos
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.utils.normalization import normalize_text, calcular_atingimento
from src.utils.styling import style_output_workbook, light_fill, PALETTE, match_group
from src.utils.logging import ValidationLogger


def test_normalize_text():
    """Testa a função normalize_text com diversos casos."""
    print("\n=== Testando normalize_text() ===")
    
    # Teste 1: String normal com acentos
    resultado = normalize_text("José da Silva")
    assert resultado == "JOSE DA SILVA", f"Esperado 'JOSE DA SILVA', obtido '{resultado}'"
    print("[OK] Teste 1: String com acentos")
    
    # Teste 2: String com BOM
    resultado = normalize_text("\ufeffTexto com BOM")
    assert resultado == "TEXTO COM BOM", f"Esperado 'TEXTO COM BOM', obtido '{resultado}'"
    print("[OK] Teste 2: String com BOM")
    
    # Teste 3: Valor NaN
    resultado = normalize_text(pd.NA)
    assert resultado == "", f"Esperado '', obtido '{resultado}'"
    print("[OK] Teste 3: Valor NaN")
    
    # Teste 4: String com espaços extras
    resultado = normalize_text("  texto   com   espaços  ")
    assert resultado == "TEXTO COM ESPACOS", f"Esperado 'TEXTO COM ESPACOS', obtido '{resultado}'"
    print("[OK] Teste 4: String com espaços extras")
    
    # Teste 5: String vazia
    resultado = normalize_text("")
    assert resultado == "", f"Esperado '', obtido '{resultado}'"
    print("[OK] Teste 5: String vazia")
    
    print("[OK] Todos os testes de normalize_text() passaram!\n")


def test_calcular_atingimento():
    """Testa a função calcular_atingimento com diversos casos."""
    print("\n=== Testando calcular_atingimento() ===")
    
    # Teste 1: Meta positiva, realizado menor
    resultado = calcular_atingimento(50, 100)
    assert resultado == 0.5, f"Esperado 0.5, obtido {resultado}"
    print("[OK] Teste 1: Meta positiva, realizado menor")
    
    # Teste 2: Meta positiva, realizado maior
    resultado = calcular_atingimento(150, 100)
    assert resultado == 1.5, f"Esperado 1.5, obtido {resultado}"
    print("[OK] Teste 2: Meta positiva, realizado maior")
    
    # Teste 3: Meta zero, realizado zero
    resultado = calcular_atingimento(0, 0)
    assert resultado == 0.0, f"Esperado 0.0, obtido {resultado}"
    print("[OK] Teste 3: Meta zero, realizado zero")
    
    # Teste 4: Meta zero, realizado positivo
    resultado = calcular_atingimento(10, 0)
    assert resultado == 1.0, f"Esperado 1.0, obtido {resultado}"
    print("[OK] Teste 4: Meta zero, realizado positivo")
    
    # Teste 5: Valores None
    resultado = calcular_atingimento(None, 100)
    assert resultado == 0.0, f"Esperado 0.0, obtido {resultado}"
    print("[OK] Teste 5: Valores None")
    
    # Teste 6: Meta zero, realizado None
    resultado = calcular_atingimento(None, 0)
    assert resultado == 0.0, f"Esperado 0.0, obtido {resultado}"
    print("[OK] Teste 6: Meta zero, realizado None")
    
    print("[OK] Todos os testes de calcular_atingimento() passaram!\n")


def test_styling_functions():
    """Testa as funções de estilização."""
    print("\n=== Testando funções de estilização ===")
    
    # Teste 1: light_fill
    fill = light_fill("E3F2FD")
    assert isinstance(fill, PatternFill), "light_fill deve retornar um PatternFill"
    # Verificar se a cor foi definida (openpyxl retorna Color object com rgb='00E3F2FD')
    assert hasattr(fill.start_color, 'rgb'), "start_color deve ter atributo rgb"
    assert fill.start_color.rgb == "00E3F2FD" or fill.start_color.rgb.endswith("E3F2FD"), f"Cor deve ser E3F2FD, obtido rgb={fill.start_color.rgb}"
    print("[OK] Teste 1: light_fill()")
    
    # Teste 2: PALETTE
    assert isinstance(PALETTE, list), "PALETTE deve ser uma lista"
    assert len(PALETTE) > 0, "PALETTE não deve estar vazia"
    print("[OK] Teste 2: PALETTE")
    
    # Teste 3: match_group
    resultado = match_group("PESO_FAT_LINHA")
    assert resultado == "faturamento_linha", f"Esperado 'faturamento_linha', obtido '{resultado}'"
    print("[OK] Teste 3: match_group()")
    
    # Teste 4: match_group com padrão não encontrado
    resultado = match_group("COLUNA_INEXISTENTE")
    assert resultado is None, f"Esperado None, obtido '{resultado}'"
    print("[OK] Teste 4: match_group() com padrão não encontrado")
    
    print("[OK] Todos os testes de estilização passaram!\n")


def test_style_output_workbook():
    """Testa a função style_output_workbook com um arquivo Excel de exemplo."""
    print("\n=== Testando style_output_workbook() ===")
    
    # Criar um arquivo Excel de teste
    wb = Workbook()
    ws = wb.active
    ws.title = "COMISSOES_CALCULADAS"
    
    # Adicionar cabeçalhos que devem ser coloridos
    ws["A1"] = "PESO_FAT_LINHA"
    ws["B1"] = "META_FAT_LINHA"
    ws["C1"] = "REALIZADO_FAT_LINHA"
    ws["D1"] = "PESO_CONV_LINHA"
    ws["E1"] = "META_CONV_LINHA"
    ws["F1"] = "Outra Coluna"
    
    # Adicionar alguns dados
    for i in range(2, 5):
        ws[f"A{i}"] = 0.3
        ws[f"B{i}"] = 1000
        ws[f"C{i}"] = 800
    
    test_file = "test_output_styling.xlsx"
    wb.save(test_file)
    
    try:
        # Aplicar estilização
        style_output_workbook(test_file)
        
        # Verificar se o arquivo ainda existe e pode ser aberto
        from openpyxl import load_workbook
        wb_test = load_workbook(test_file)
        assert "COMISSOES_CALCULADAS" in wb_test.sheetnames, "Aba deve existir"
        print("[OK] Teste: style_output_workbook() aplicou estilização com sucesso")
        
        wb_test.close()
    except Exception as e:
        print(f"[ERRO] Erro ao testar style_output_workbook(): {e}")
        raise
    finally:
        # Limpar arquivo de teste
        if os.path.exists(test_file):
            os.remove(test_file)
    
    print("[OK] Todos os testes de style_output_workbook() passaram!\n")


def test_validation_logger():
    """Testa a classe ValidationLogger."""
    print("\n=== Testando ValidationLogger ===")
    
    logger = ValidationLogger()
    
    # Teste 1: Adicionar logs
    logger.info("Mensagem de info")
    logger.aviso("Mensagem de aviso", {"campo": "valor"})
    logger.erro("Mensagem de erro")
    
    assert len(logger) == 3, f"Esperado 3 logs, obtido {len(logger)}"
    print("[OK] Teste 1: Adicionar logs")
    
    # Teste 2: Verificar estrutura dos logs
    logs = logger.get_logs()
    assert len(logs) == 3, "Deve haver 3 logs"
    assert logs[0]["Nível"] == "INFO", "Primeiro log deve ser INFO"
    assert logs[1]["Nível"] == "AVISO", "Segundo log deve ser AVISO"
    assert logs[2]["Nível"] == "ERRO", "Terceiro log deve ser ERRO"
    print("[OK] Teste 2: Estrutura dos logs")
    
    # Teste 3: Limpar logs
    logger.clear()
    assert len(logger) == 0, "Logs devem estar vazios após clear()"
    print("[OK] Teste 3: Limpar logs")
    
    print("[OK] Todos os testes de ValidationLogger passaram!\n")


def main():
    """Executa todos os testes."""
    print("=" * 60)
    print("TESTES DE VALIDAÇÃO DAS FUNÇÕES UTILITÁRIAS MIGRADAS")
    print("=" * 60)
    
    try:
        test_normalize_text()
        test_calcular_atingimento()
        test_styling_functions()
        test_style_output_workbook()
        test_validation_logger()
        
        print("=" * 60)
        print("[SUCESSO] TODOS OS TESTES PASSARAM COM SUCESSO!")
        print("=" * 60)
        return 0
    except AssertionError as e:
        print(f"\n[FALHA] FALHA NO TESTE: {e}")
        return 1
    except Exception as e:
        print(f"\n[ERRO] ERRO INESPERADO: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())

