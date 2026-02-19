"""
Sistema de Auditoria de Testes — Gerador de Relatório Excel.

Gera automaticamente após cada rodada de pytest um relatório Excel
em tests_comissoes/relatorios/ com todas as verificações detalhadas.

Uso:
    Cada teste registra suas verificações via fixture `audit`:

        def test_fc_escada_basico(audit):
            audit.set_contexto(modulo="FC Escada", cenario="Performance 80%")
            audit.registrar(
                descricao="Multiplicador do degrau 2 com piso 50%",
                formula="piso + (i × (1-piso) / (n-1)) = 0.5 + (2 × 0.5 / 3)",
                entradas={"performance": 0.8, "num_degraus": 4, "piso": 0.5},
                esperado=0.833,
                real=resultado,
            )

    Após o pytest finalizar, conftest.py gera o Excel automaticamente.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# =========================================================================
# Dataclass de uma verificação individual
# =========================================================================
@dataclass
class VerificacaoAuditoria:
    """Uma verificação (assertion) individual registrada pelo teste."""

    modulo: str
    cenario: str
    descricao: str
    formula: str
    entradas: Dict[str, Any]
    esperado: Any
    real: Any
    tolerancia: float = 0.01
    passou: Optional[bool] = None
    observacao: str = ""

    def __post_init__(self):
        if self.passou is None:
            self.passou = self._avaliar()

    def _avaliar(self) -> bool:
        """Avalia se o resultado real está dentro da tolerância."""
        try:
            e = float(self.esperado)
            r = float(self.real)
            return abs(e - r) <= self.tolerancia
        except (TypeError, ValueError):
            return str(self.esperado) == str(self.real)

    @property
    def diferenca(self) -> str:
        try:
            e = float(self.esperado)
            r = float(self.real)
            return f"{r - e:+.6f}"
        except (TypeError, ValueError):
            return "N/A"


# =========================================================================
# Coletor de verificações (usado como fixture)
# =========================================================================
class AuditCollector:
    """Coletor de verificações para um único teste.

    Usado via fixture pytest `audit`.
    """

    def __init__(self, test_name: str):
        self.test_name = test_name
        self.verificacoes: List[VerificacaoAuditoria] = []
        self._modulo = ""
        self._cenario = ""

    def set_contexto(self, modulo: str = "", cenario: str = ""):
        """Define módulo e cenário para verificações subsequentes."""
        if modulo:
            self._modulo = modulo
        if cenario:
            self._cenario = cenario

    def registrar(
        self,
        descricao: str,
        formula: str,
        entradas: Dict[str, Any],
        esperado: Any,
        real: Any,
        tolerancia: float = 0.01,
        observacao: str = "",
    ) -> VerificacaoAuditoria:
        """Registra uma verificação e retorna o objeto para consulta."""
        v = VerificacaoAuditoria(
            modulo=self._modulo or "Geral",
            cenario=self._cenario or self.test_name,
            descricao=descricao,
            formula=formula,
            entradas=entradas,
            esperado=esperado,
            real=real,
            tolerancia=tolerancia,
            observacao=observacao,
        )
        self.verificacoes.append(v)
        return v

    def verificar(
        self,
        descricao: str,
        formula: str,
        entradas: Dict[str, Any],
        esperado: Any,
        real: Any,
        tolerancia: float = 0.01,
        observacao: str = "",
    ):
        """Registra E faz assert. Falha o teste se resultado não bater."""
        v = self.registrar(
            descricao=descricao,
            formula=formula,
            entradas=entradas,
            esperado=esperado,
            real=real,
            tolerancia=tolerancia,
            observacao=observacao,
        )
        if not v.passou:
            raise AssertionError(
                f"FALHOU: {descricao}\n"
                f"  Esperado: {esperado}\n"
                f"  Real:     {real}\n"
                f"  Fórmula:  {formula}\n"
                f"  Diferença: {v.diferenca}"
            )


# =========================================================================
# Gerador do relatório Excel
# =========================================================================
class AuditReportGenerator:
    """Gera o relatório Excel de auditoria com todas as verificações."""

    # Cores
    _GREEN = "C6EFCE"
    _RED = "FFC7CE"
    _HEADER_BG = "4472C4"
    _HEADER_FG = "FFFFFF"

    def __init__(self):
        self.todas_verificacoes: List[VerificacaoAuditoria] = []

    def adicionar(self, verificacoes: List[VerificacaoAuditoria]):
        """Adiciona verificações de um teste."""
        self.todas_verificacoes.extend(verificacoes)

    def gerar_relatorio(self, output_dir: str) -> str:
        """Gera o relatório Excel e retorna o caminho do arquivo."""
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filepath = os.path.join(output_dir, f"auditoria_testes_{timestamp}.xlsx")

        # Construir DataFrames
        df_resumo = self._construir_resumo()
        df_detalhe = self._construir_detalhe()
        dfs_por_modulo = self._construir_por_modulo()

        # Escrever Excel
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_resumo.to_excel(writer, sheet_name="RESUMO", index=False)
            df_detalhe.to_excel(writer, sheet_name="TODAS_VERIFICACOES", index=False)

            for modulo, df_mod in dfs_por_modulo.items():
                # Nome de aba com no máximo 31 caracteres
                sheet_name = modulo[:31].replace("/", "_")
                df_mod.to_excel(writer, sheet_name=sheet_name, index=False)

            # Aplicar estilos
            if _HAS_OPENPYXL:
                self._aplicar_estilos(writer)

        return filepath

    # ------------------------------------------------------------------
    # Construção dos DataFrames
    # ------------------------------------------------------------------
    def _construir_resumo(self) -> pd.DataFrame:
        """Resumo por módulo: total, passou, falhou, % aprovação."""
        modulos = {}
        for v in self.todas_verificacoes:
            if v.modulo not in modulos:
                modulos[v.modulo] = {"total": 0, "passou": 0, "falhou": 0}
            modulos[v.modulo]["total"] += 1
            if v.passou:
                modulos[v.modulo]["passou"] += 1
            else:
                modulos[v.modulo]["falhou"] += 1

        rows = []
        total_geral = {"Módulo": "TOTAL GERAL", "Total": 0, "Passou ✅": 0, "Falhou ❌": 0}
        for mod, counts in sorted(modulos.items()):
            pct = (counts["passou"] / counts["total"] * 100) if counts["total"] > 0 else 0
            rows.append({
                "Módulo": mod,
                "Total": counts["total"],
                "Passou ✅": counts["passou"],
                "Falhou ❌": counts["falhou"],
                "Aprovação %": f"{pct:.1f}%",
            })
            total_geral["Total"] += counts["total"]
            total_geral["Passou ✅"] += counts["passou"]
            total_geral["Falhou ❌"] += counts["falhou"]

        pct_geral = (total_geral["Passou ✅"] / total_geral["Total"] * 100) if total_geral["Total"] > 0 else 0
        total_geral["Aprovação %"] = f"{pct_geral:.1f}%"
        rows.append(total_geral)

        return pd.DataFrame(rows)

    def _construir_detalhe(self) -> pd.DataFrame:
        """Todas as verificações em uma única tabela."""
        rows = []
        for v in self.todas_verificacoes:
            rows.append({
                "Status": "✅ PASSOU" if v.passou else "❌ FALHOU",
                "Módulo": v.modulo,
                "Cenário": v.cenario,
                "Verificação": v.descricao,
                "Fórmula": v.formula,
                "Entradas": str(v.entradas),
                "Esperado": v.esperado,
                "Real": v.real,
                "Diferença": v.diferenca,
                "Tolerância": v.tolerancia,
                "Observação": v.observacao,
            })
        return pd.DataFrame(rows)

    def _construir_por_modulo(self) -> Dict[str, pd.DataFrame]:
        """Um DataFrame por módulo."""
        por_modulo: Dict[str, list] = {}
        for v in self.todas_verificacoes:
            if v.modulo not in por_modulo:
                por_modulo[v.modulo] = []
            por_modulo[v.modulo].append({
                "Status": "✅ PASSOU" if v.passou else "❌ FALHOU",
                "Cenário": v.cenario,
                "Verificação": v.descricao,
                "Fórmula Aplicada": v.formula,
                "Entradas": str(v.entradas),
                "Resultado Esperado": v.esperado,
                "Resultado Real": v.real,
                "Diferença": v.diferenca,
                "Observação": v.observacao,
            })
        return {mod: pd.DataFrame(rows) for mod, rows in por_modulo.items()}

    # ------------------------------------------------------------------
    # Estilos visuais
    # ------------------------------------------------------------------
    def _aplicar_estilos(self, writer: pd.ExcelWriter):
        """Aplica cores, fontes e largura de colunas em todas as abas."""
        wb = writer.book

        header_fill = PatternFill(start_color=self._HEADER_BG, end_color=self._HEADER_BG, fill_type="solid")
        header_font = Font(color=self._HEADER_FG, bold=True, size=11)
        passed_fill = PatternFill(start_color=self._GREEN, end_color=self._GREEN, fill_type="solid")
        failed_fill = PatternFill(start_color=self._RED, end_color=self._RED, fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for ws in wb.worksheets:
            # Estilizar cabeçalho
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Largura automática
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col_cells:
                    try:
                        cell_len = len(str(cell.value or ""))
                        max_len = max(max_len, cell_len)
                    except Exception:
                        pass
                    cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

            # Colorir linhas de status
            status_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value and "Status" in str(cell.value):
                    status_col = col_idx
                    break

            if status_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    status_cell = row[status_col - 1]
                    status_val = str(status_cell.value or "")
                    if "PASSOU" in status_val:
                        for cell in row:
                            cell.fill = passed_fill
                    elif "FALHOU" in status_val:
                        for cell in row:
                            cell.fill = failed_fill
